import random
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY


pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
pdfmetrics.registerFont(TTFont('Arial-Italic', 'ArialItalic.ttf'))


class GenerateExamTickets:

    def __init__(self, config):
        self.num_tickets = config["num_tickets"]
        self.questions_file = config["exam_questions_file"]
        self.output_pdf = config["output_pdf"]

        self.text = config["text"]

    def generate(self):
        try:
            part1, part2 = self.read_questions(self.questions_file)
            tickets = self.generate_unique_tickets(part1, part2, self.num_tickets)
            self.create_pdf(tickets, self.output_pdf)
            print(f"Document has been created successfully: {self.output_pdf}")
        except Exception as e:
            print(f"Error: {e}")

    @staticmethod
    def read_questions(path):
        wb = load_workbook(path)
        sheet = wb.active
        part1, part2 = [], []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            q1, q2 = row
            if q1:
                part1.append(str(q1).strip())
            if q2:
                part2.append(str(q2).strip())
        return part1, part2

    @staticmethod
    def generate_unique_tickets(part1, part2, count):
        generated = set()
        tickets = []

        while len(tickets) < count:
            q1 = random.choice(part1)
            q2 = random.choice(part2)
            key = (q1, q2)

            if key not in generated:
                tickets.append(key)
                generated.add(key)
            else:
                for _ in range(10):
                    if random.random() < 0.5:
                        q1 = random.choice(part1)
                    else:
                        q2 = random.choice(part2)
                    key = (q1, q2)
                    if key not in generated:
                        tickets.append(key)
                        generated.add(key)
                        break
                else:
                    tickets.append(key)
        return tickets

    @staticmethod
    def draw_wrapped_paragraph(c, text, x, y, max_width, font_name="Arial", font_size=11, leading=14):
        text = text.replace('\n', '<br/>')
        style = ParagraphStyle(
            name='Justify',
            fontName=font_name,
            fontSize=font_size,
            leading=leading,
            alignment=TA_JUSTIFY
        )
        p = Paragraph(text, style)
        w, h = p.wrap(max_width, 1000)
        p.drawOn(c, x, y - h)
        return h

    def create_pdf(self, tickets, output_path):
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        margin = 2 * cm
        usable_width = width - 2 * margin

        def draw_ticket(ticket_num, q1, q2, y_offset, bottom):
            y = y_offset
            c.setFont("Arial", 10)
            c.drawCentredString(width / 2, y, self.text["exam_university"])
            y -= 5
            c.line(margin, y, width - margin, y)
            y -= 20

            c.setFont("Arial-Bold", 16)
            c.drawCentredString(width / 2, y, self.text["exam_header_title"] + f" №{ticket_num}")
            y -= 20

            c.setFont("Arial-Italic", 12)
            c.drawString(margin, y, self.text["exam_ticket_title"] + self.text["exam_subject"])
            y -= 15
            c.drawString(margin, y, self.text["exam_info"])
            y -= 5
            c.line(margin, y, width - margin, y)
            y -= 25

            h1 = self.draw_wrapped_paragraph(c, f"1. {q1}", margin, y, usable_width)
            y -= h1 + 12

            h2 = self.draw_wrapped_paragraph(c, f"2. {q2}", margin, y, usable_width)
            y -= h2 + 25

            c.setFont("Arial", 12)
            c.drawRightString(width - margin, bottom + margin,
                              self.text["exam_teacher_title"] + self.text["exam_teacher"])

        for i in range(0, len(tickets), 2):
            ticket1 = tickets[i]
            ticket2 = tickets[i + 1] if i + 1 < len(tickets) else ("", "")
            draw_ticket(i + 1, *ticket1, y_offset=height - margin, bottom=height / 2)
            c.line(0, height / 2, width, height / 2)
            draw_ticket(i + 2, *ticket2, y_offset=height / 2 - margin, bottom=0)
            c.showPage()

        c.save()
